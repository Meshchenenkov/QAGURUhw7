import os
import zipfile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook


CURRENT_FILE = os.path.abspath(__file__) # путь к файлу
CURRENT_DIRECTORY = os.path.dirname(CURRENT_FILE) # по файлу находим путь к папке, где он лежит
RESOURCES_DIRECTORY = os.path.join(CURRENT_DIRECTORY, 'resources') # путь к папке resources
FILES_DIRECTORY = os.path.join(CURRENT_DIRECTORY, "files") # путь к папке tmp


def test_read_csv():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIRECTORY, 'zip.zip')) as zf:
        with zf.open('csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            readcsv = list(csv.reader(content.splitlines())) #склеиваем содержимое строк в список
            assert 'Hashimoto' in str(readcsv[2])


def test_read_pdf():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIRECTORY, 'zip.zip')) as zf:
        with zf.open('pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            assert 'Тестовый PDF файл' in reader.pages[0].extract_text()


def test_read_xlsx():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIRECTORY, 'zip.zip')) as zf:
        with zf.open('xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            for_assert = [[0, 'First Name', 'Last Name', 'Gender', 'Country', 'Age', 'Date', 'Id'],
                          [1, 'Dulce', 'Abril', 'Female', 'United States', 32, '15/10/2017', 1562],
                          [2, 'Mara', 'Hashimoto', 'Female', 'Great Britain', 25, '16/08/2016', 1582],
                          [3, 'Philip', 'Gent', 'Male', 'France', 36, '21/05/2015', 2587],
                          [4, 'Kathleen', 'Hanner', 'Female', 'United States', 25, '15/10/2017', 3549],
                          [5, 'Nereida', 'Magwood', 'Female', 'United States', 58, '16/08/2016', 2468],
                          [6, 'Gaston', 'Brumm', 'Male', 'United States', 24, '21/05/2015', 2554],
                          [7, 'Etta', 'Hurn', 'Female', 'Great Britain', 56, '15/10/2017', 3598],
                          [8, 'Earlean', 'Melgar', 'Female', 'United States', 27, '16/08/2016', 2456],
                          [9, 'Vincenza', 'Weiland', 'Female', 'United States', 40, '21/05/2015', 6548],
                          [10, 'Fallon', 'Winward', 'Female', 'Great Britain', 28, '16/08/2016', 5486],
                          [11, 'Arcelia', 'Bouska', 'Female', 'Great Britain', 39, '21/05/2015', 1258],
                          [12, 'Franklyn', 'Unknow', 'Male', 'France', 38, '15/10/2017', 2579],
                          [13, 'Sherron', 'Ascencio', 'Female', 'Great Britain', 32, '16/08/2016', 3256],
                          [14, 'Marcel', 'Zabriskie', 'Male', 'Great Britain', 26, '21/05/2015', 2587],
                          [15, 'Kina', 'Hazelton', 'Female', 'Great Britain', 31, '16/08/2016', 3259],
                          [16, 'Shavonne', 'Pia', 'Female', 'France', 24, '21/05/2015', 1546],
                          [17, 'Shavon', 'Benito', 'Female', 'France', 39, '15/10/2017', 3579],
                          [18, 'Lauralee', 'Perrine', 'Female', 'Great Britain', 28, '16/08/2016', 6597],
                          [19, 'Loreta', 'Curren', 'Female', 'France', 26, '21/05/2015', 9654],
                          [20, 'Teresa', 'Strawn', 'Female', 'France', 46, '21/05/2015', 3569],
                          [21, 'Belinda', 'Partain', 'Female', 'United States', 37, '15/10/2017', 2564],
                          [22, 'Holly', 'Eudy', 'Female', 'United States', 52, '16/08/2016', 8561],
                          [23, 'Many', 'Cuccia', 'Female', 'Great Britain', 46, '21/05/2015', 5489],
                          [24, 'Libbie', 'Dalby', 'Female', 'France', 42, '21/05/2015', 5489],
                          [25, 'Lester', 'Prothro', 'Male', 'France', 21, '15/10/2017', 6574],
                          [26, 'Marvel', 'Hail', 'Female', 'Great Britain', 28, '16/08/2016', 5555],
                          [27, 'Angelyn', 'Vong', 'Female', 'United States', 29, '21/05/2015', 6125],
                          [28, 'Francesca', 'Beaudreau', 'Female', 'France', 23, '15/10/2017', 5412],
                          [29, 'Garth', 'Gangi', 'Male', 'United States', 41, '16/08/2016', 3256],
                          [30, 'Carla', 'Trumbull', 'Female', 'Great Britain', 28, '21/05/2015', 3264],
                          [31, 'Veta', 'Muntz', 'Female', 'Great Britain', 37, '15/10/2017', 4569],
                          [32, 'Stasia', 'Becker', 'Female', 'Great Britain', 34, '16/08/2016', 7521],
                          [33, 'Jona', 'Grindle', 'Female', 'Great Britain', 26, '21/05/2015', 6458],
                          [34, 'Judie', 'Claywell', 'Female', 'France', 35, '16/08/2016', 7569],
                          [35, 'Dewitt', 'Borger', 'Male', 'United States', 36, '21/05/2015', 8514],
                          [36, 'Nena', 'Hacker', 'Female', 'United States', 29, '15/10/2017', 8563],
                          [37, 'Kelsie', 'Wachtel', 'Female', 'France', 27, '16/08/2016', 8642],
                          [38, 'Sau', 'Pfau', 'Female', 'United States', 25, '21/05/2015', 9536],
                          [39, 'Shanice', 'Mccrystal', 'Female', 'United States', 36, '21/05/2015', 2567],
                          [40, 'Chase', 'Karner', 'Male', 'United States', 37, '15/10/2017', 2154],
                          [41, 'Tommie', 'Underdahl', 'Male', 'United States', 26, '16/08/2016', 3265],
                          [42, 'Dorcas', 'Darity', 'Female', 'United States', 37, '21/05/2015', 8765],
                          [43, 'Angel', 'Sanor', 'Male', 'France', 24, '15/10/2017', 3259],
                          [44, 'Willodean', 'Harn', 'Female', 'United States', 39, '16/08/2016', 3567],
                          [45, 'Weston', 'Martina', 'Male', 'United States', 26, '21/05/2015', 6540],
                          [46, 'Roma', 'Lafollette', 'Female', 'United States', 34, '15/10/2017', 2654],
                          [47, 'Felisa', 'Cail', 'Female', 'United States', 28, '16/08/2016', 6525],
                          [48, 'Demetria', 'Abbey', 'Female', 'United States', 32, '21/05/2015', 3265],
                          [49, 'Jeromy', 'Danz', 'Male', 'United States', 39, '15/10/2017', 3265],
                          [50, 'Rasheeda', 'Alkire', 'Female', 'United States', 29, '16/08/2016', 6125]]
            assert data == for_assert






